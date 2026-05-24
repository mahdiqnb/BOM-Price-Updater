# -*- coding: utf-8 -*-
"""
BOM PRICE UPDATER - PRECISION EDITION v9.0
دقیق‌ترین کد با سلکتورهای تخصصی هر سایت
✓ ICKALA | ECA | SKYTECH | LION ELECTRONIC
"""

import sys, re, time, shutil, threading, json
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

def _install(pkg, imp=None):
    try:
        __import__(imp or pkg)
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg],
                              stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

for _p, _i in [("requests",None), ("beautifulsoup4","bs4"),
                ("openpyxl",None), ("lxml",None)]:
    _install(_p, _i)

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "fa-IR,fa;q=0.9,en-US;q=0.8",
})

FILL_GREEN  = PatternFill(fill_type="solid", start_color="C6EFCE")
FILL_RED    = PatternFill(fill_type="solid", start_color="FFC7CE")
FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFEB9C")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PRICE CLEANING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def clean_price(text):
    """تمیزکاری قیمت"""
    if not text:
        return None
    text = str(text).strip()
    text = text.translate(str.maketrans(
        "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789"))
    for r in [",","٬","،","ریال","تومان","قیمت","IRR","﷼"," ","\n","\r","\t","\u200c","\xa0","<small>","</small>"]:
        text = text.replace(r, "")
    m = re.search(r'(\d{3,})', text)
    if not m:
        return None
    try:
        val = int(m.group(1))
        if 100 <= val < 100000000:
            return val
    except:
        pass
    return None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SITE-SPECIFIC PARSERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def parse_ickala(soup, html):
    """ICKALA Parser"""
    avail_tag = soup.select_one("span#availability_value.warning_inline")
    if avail_tag and "موجود نیست" in avail_tag.get_text():
        return None, "ناموجود"
    
    price_tag = soup.select_one("span#our_price_display.price_display_originalvalue")
    if price_tag:
        p = clean_price(price_tag.get_text())
        if p:
            return p, "OK"
    
    for sel in ["span#our_price_display", ".price_display_originalvalue", ".price"]:
        tag = soup.select_one(sel)
        if tag:
            p = clean_price(tag.get_text())
            if p:
                return p, "OK"
    
    return None, "PRICE NOT FOUND"

def parse_lion(soup, html):
    """LION ELECTRONIC Parser"""
    stock_tag = soup.select_one("span.stock-amount")
    if stock_tag:
        stock_text = stock_tag.get_text().strip()
        if "اتمام موجودی" in stock_text or "تمام شده" in stock_text:
            return None, "ناموجود"
    
    price_tag = soup.select_one("span.new-price")
    if price_tag:
        p = clean_price(price_tag.get_text())
        if p:
            return p, "OK"
    
    for sel in ["span.price", ".price-group", ".woocommerce-Price-amount"]:
        tag = soup.select_one(sel)
        if tag:
            p = clean_price(tag.get_text())
            if p:
                return p, "OK"
    
    return None, "PRICE NOT FOUND"

def parse_eca(soup, html):
    """ECA Parser"""
    outstock = soup.select_one("span.out-of-stock")
    if outstock:
        return None, "ناموجود"
    
    price_tag = soup.select_one("span.current-price.fa-number-conv")
    if price_tag:
        p = clean_price(price_tag.get_text())
        if p:
            return p, "OK"
    
    for sel in [".current-price", ".price", "span.price"]:
        tag = soup.select_one(sel)
        if tag:
            p = clean_price(tag.get_text())
            if p:
                return p, "OK"
    
    return None, "PRICE NOT FOUND"

def parse_skytech(soup, html):
    """SKYTECH Parser"""
    label9 = soup.select_one("span#Label9")
    if label9:
        text = label9.get_text().strip()
        if "تمام شده" in text or "موجودی" not in text and not text.isdigit():
            return None, "ناموجود"
    
    price_tag = soup.select_one("span#DataList21_Label1_0")
    if not price_tag:
        price_tag = soup.select_one("span.price")
    
    if price_tag:
        p = clean_price(price_tag.get_text())
        if p:
            return p, "OK"
    
    prod_price = soup.select_one("div.prod_price")
    if prod_price:
        p = clean_price(prod_price.get_text())
        if p:
            return p, "OK"
    
    return None, "PRICE NOT FOUND"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN FETCHER
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def fetch_price(url):
    """Fetch price from URL using site-specific parser"""
    try:
        url = url.strip()
        r = session.get(url, timeout=20)
        
        if r.status_code != 200:
            return None, f"HTTP {r.status_code}"
        
        soup = BeautifulSoup(r.text, "lxml")
        domain = url.lower()
        
        if "ickala.com" in domain:
            price, status = parse_ickala(soup, r.text)
        elif "lionelectronic.ir" in domain or "lion-electronic" in domain:
            price, status = parse_lion(soup, r.text)
        elif "eca.ir" in domain or "eshop.eca.ir" in domain:
            price, status = parse_eca(soup, r.text)
        elif "skytech.ir" in domain:
            price, status = parse_skytech(soup, r.text)
        else:
            price = None
            status = "UNKNOWN SITE"
            for sel in [".price", "[class*='price']", "span.price"]:
                tag = soup.select_one(sel)
                if tag:
                    p = clean_price(tag.get_text())
                    if p:
                        price, status = p, "OK"
                        break
            if not price:
                status = "PRICE NOT FOUND"
        
        return price, status
    
    except Exception as e:
        return None, str(e)[:30]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROCESS EXCEL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def process_excel(path, log_fn, progress_fn, finish_fn):
    try:
        src = Path(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = src.with_name(src.stem + f"_BACKUP_{ts}" + src.suffix)
        shutil.copy2(src, backup)

        wb = load_workbook(path)
        ws = wb.active

        LINK_COL = 7
        QTY_COL  = 4
        PRICE_COL = ws.max_column + 1
        TOTAL_COL = PRICE_COL + 1

        for col, text in [(PRICE_COL,"UNIT PRICE"), (TOTAL_COL,"TOTAL")]:
            c = ws.cell(1, col)
            c.value = text
            c.fill = PatternFill(fill_type="solid", start_color="1a3a5c")
            c.font = Font(color="FFFFFF", bold=True)
            c.alignment = Alignment(horizontal="center")

        ok = fail = err = 0
        total_rows = ws.max_row - 1

        for row in range(2, ws.max_row + 1):
            try:
                cell = ws.cell(row, LINK_COL)
                url = None
                
                if cell.hyperlink:
                    url = cell.hyperlink.target
                if not url:
                    url = str(cell.value or "").strip()

                if not url or not url.startswith("http"):
                    ws.cell(row, PRICE_COL).value = "INVALID"
                    for c in range(1, TOTAL_COL+1):
                        ws.cell(row, c).fill = FILL_YELLOW
                    err += 1
                    progress_fn(row-1, total_rows, ok, fail, err)
                    log_fn(f"ROW {row:02d}  ❌  INVALID LINK", "error")
                    continue

                qty = 0
                try:
                    qty = int(ws.cell(row, QTY_COL).value or 0)
                except:
                    qty = 0

                domain = url.split('/')[2] if '//' in url else url[:30]
                log_fn(f"ROW {row:02d}  →  {domain}", "info")
                
                price, status = fetch_price(url)

                if price and price > 100:
                    total = price * qty
                    ws.cell(row, PRICE_COL).value = price
                    ws.cell(row, PRICE_COL).number_format = '#,##0'
                    ws.cell(row, TOTAL_COL).value = total
                    ws.cell(row, TOTAL_COL).number_format = '#,##0'
                    for c in range(1, TOTAL_COL+1):
                        ws.cell(row, c).fill = FILL_GREEN
                    ok += 1
                    log_fn(f"ROW {row:02d}  ✓  {price:,}", "success")

                elif status == "ناموجود":
                    ws.cell(row, PRICE_COL).value = "ناموجود"
                    for c in range(1, TOTAL_COL+1):
                        ws.cell(row, c).fill = FILL_RED
                    fail += 1
                    log_fn(f"ROW {row:02d}  ✗  UNAVAILABLE", "fail")

                else:
                    ws.cell(row, PRICE_COL).value = status
                    for c in range(1, TOTAL_COL+1):
                        ws.cell(row, c).fill = FILL_YELLOW
                    err += 1
                    log_fn(f"ROW {row:02d}  ⚠  {status}", "warning")

                progress_fn(row-1, total_rows, ok, fail, err)
                time.sleep(0.8)

            except Exception as e:
                err += 1
                log_fn(f"ROW {row:02d}  ❌  Exception", "error")
                progress_fn(row-1, total_rows, ok, fail, err)

        out_dir = Path("UPDATED_BOMS")
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"{src.stem}_UPDATED_{ts}.xlsx"
        wb.save(out_path)
        finish_fn(out_path, ok, fail, err)

    except Exception as e:
        finish_fn(None, 0, 0, 0, str(e))


# ════════════════════════════════════════════════════════
#                    GUI v9.0+
# ════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BOM Price Updater — PRECISION v9.0")
        self.geometry("1400x950")
        self.minsize(1100, 800)
        
        self.bg_main = "#0f1419"
        self.bg_panel = "#1a1f2e"
        self.bg_input = "#232d3f"
        self.accent = "#00d9ff"
        self.success = "#00ff9f"
        self.danger = "#ff4757"
        self.warning = "#ffa502"
        self.text_main = "#f5f7fa"
        self.text_sub = "#b8bec7"
        
        self.configure(bg=self.bg_main)
        self._file_path = tk.StringVar()
        self._running = False
        
        self._build_ui()

    def _build_ui(self):
        # ━━━ HEADER ━━━
        header = tk.Frame(self, bg=self.bg_main)
        header.pack(fill="x", padx=40, pady=16)
        
        tk.Label(header, text="◆ BOM PRICE UPDATER — PRECISION v9.0",
                font=("Courier New", 24, "bold"),
                fg=self.accent, bg=self.bg_main).pack(anchor="w")
        
        tk.Label(header, text="دقیق‌ترین کد با سلکتورهای تخصصی هر سایت",
                font=("Courier New", 9),
                fg=self.text_sub, bg=self.bg_main).pack(anchor="w")
        
        # ━━━ SUPPORTED SITES ━━━
        sites_frame = tk.Frame(header, bg=self.bg_main)
        sites_frame.pack(fill="x", pady=(12, 0))
        
        tk.Label(sites_frame, text="✓ SUPPORTED SITES:",
                font=("Courier New", 8, "bold"),
                fg=self.text_sub, bg=self.bg_main).pack(anchor="w")
        
        sites_row = tk.Frame(sites_frame, bg=self.bg_main)
        sites_row.pack(anchor="w", pady=(4, 0))
        
        sites = [("ICKALA", self.accent), ("ECA", self.success), 
                 ("SKYTECH", self.warning), ("LION ELECTRONIC", self.danger)]
        for site, color in sites:
            f = tk.Frame(sites_row, bg=color, padx=8, pady=2)
            f.pack(side="left", padx=4)
            tk.Label(f, text=site, font=("Courier New", 7, "bold"),
                    fg=self.bg_main, bg=color).pack()
        
        # ━━━ SEPARATOR ━━━
        tk.Frame(self, bg=self.bg_panel, height=2).pack(fill="x", pady=(8, 0))
        
        # ━━━ INPUT SECTION ━━━
        input_panel = tk.Frame(self, bg=self.bg_panel)
        input_panel.pack(fill="x", padx=40, pady=14)
        
        inner = tk.Frame(input_panel, bg=self.bg_panel)
        inner.pack(fill="x", padx=20, pady=14)
        
        row = tk.Frame(inner, bg=self.bg_panel)
        row.pack(fill="x")
        
        self._entry = tk.Entry(row, textvariable=self._file_path,
                              font=("Courier New", 11),
                              bg=self.bg_input, fg=self.text_main,
                              relief="flat", bd=0)
        self._entry.pack(side="left", fill="x", expand=True, ipady=10, padx=(0, 12))
        
        tk.Button(row, text="📂 BROWSE", command=self._browse,
                 font=("Courier New", 10, "bold"),
                 bg=self.bg_input, fg=self.accent,
                 relief="flat", bd=0, padx=20, pady=10,
                 activebackground=self.accent, activeforeground=self.bg_main).pack(side="left", padx=4)
        
        tk.Button(row, text="▶ START", command=self._start,
                 font=("Courier New", 10, "bold"),
                 bg=self.success, fg=self.bg_main,
                 relief="flat", bd=0, padx=20, pady=10,
                 activebackground=self.bg_main, activeforeground=self.success).pack(side="left", padx=4)
        
        # ━━━ PROGRESS BAR ━━━
        self._pbar = tk.Canvas(self, bg=self.bg_input, highlightthickness=0, height=8)
        self._pbar.pack(fill="x", padx=40, pady=(0, 12))
        self._pbar_value = 0
        self._pbar.bind("<Configure>", self._redraw_pbar)
        
        # ━━━ STATS ━━━
        stats_frame = tk.Frame(self, bg=self.bg_main)
        stats_frame.pack(fill="x", padx=40, pady=(0, 14))
        for i in range(4):
            stats_frame.columnconfigure(i, weight=1)
        
        self.stat_values = {}
        for idx, (label, color, key) in enumerate([
            ("SUCCESS", self.success, "ok"), ("UNAVAIL", self.danger, "fail"),
            ("ERRORS", self.warning, "err"), ("TOTAL", self.accent, "tot")
        ]):
            box = tk.Frame(self, bg=self.bg_panel)
            box.grid(row=0, column=idx, in_=stats_frame, sticky="nsew", padx=6, pady=0)
            
            v = tk.Label(box, text="0", font=("Courier New", 22, "bold"),
                        fg=color, bg=self.bg_panel)
            v.pack(pady=(12, 4))
            tk.Label(box, text=label, font=("Courier New", 8),
                    fg=self.text_sub, bg=self.bg_panel).pack(pady=(0, 12))
            self.stat_values[key] = v
        
        # ━━━ LOG SECTION ━━━
        log_header = tk.Frame(self, bg=self.bg_panel)
        log_header.pack(fill="x", padx=40, pady=(12, 0))
        tk.Label(log_header, text="● LIVE LOG",
                font=("Courier New", 10, "bold"),
                fg=self.accent, bg=self.bg_panel).pack(anchor="w", padx=16, pady=12)
        
        log_container = tk.Frame(self, bg=self.bg_panel)
        log_container.pack(fill="both", expand=True, padx=40, pady=(0, 20))
        
        scrollbar = tk.Scrollbar(log_container)
        scrollbar.pack(side="right", fill="y")
        
        self._log = tk.Text(log_container, font=("Courier New", 9),
                           bg=self.bg_input, fg=self.text_main,
                           relief="flat", bd=0, padx=12, pady=10,
                           yscrollcommand=scrollbar.set)
        self._log.pack(fill="both", expand=True)
        scrollbar.config(command=self._log.yview)
        
        # ━━━ LOG COLORS ━━━
        self._log.tag_configure("success", foreground=self.success, font=("Courier New", 9, "bold"))
        self._log.tag_configure("fail", foreground=self.danger, font=("Courier New", 9, "bold"))
        self._log.tag_configure("warning", foreground=self.warning, font=("Courier New", 9, "bold"))
        self._log.tag_configure("error", foreground="#ff1744", font=("Courier New", 9, "bold"))
        self._log.tag_configure("info", foreground=self.accent, font=("Courier New", 9))

    def _redraw_pbar(self, evt=None):
        w = self._pbar.winfo_width()
        if w <= 1:
            return
        self._pbar.delete("all")
        self._pbar.create_rectangle(0, 0, w, 8, fill=self.bg_input, outline="")
        fw = int(w * self._pbar_value / 100)
        if fw > 0:
            self._pbar.create_rectangle(0, 0, fw, 8, fill=self.accent, outline="")
            self._pbar.create_rectangle(max(0, fw-3), 0, fw, 8, fill=self.success, outline="")

    def _browse(self):
        p = filedialog.askopenfilename(title="Select BOM Excel File",
                                       filetypes=[("Excel", "*.xlsx")])
        if p:
            self._file_path.set(p)
            self._log_msg(f"📁 File loaded: {Path(p).name}", "info")

    def _start(self):
        if self._running:
            return
        path = self._file_path.get().strip()
        if not path or not Path(path).exists():
            messagebox.showerror("Error", "Select valid Excel file")
            return

        self._running = True
        self._pbar_value = 0
        self._log.delete("1.0", "end")
        self._log_msg("▶ Starting processing…", "info")

        def worker():
            process_excel(path,
                         lambda msg, tag: self.after(0, self._log_msg, msg, tag),
                         lambda d,t,o,f,e: self.after(0, self._update_stats, d,t,o,f,e),
                         lambda out,o,f,e,em=None: self.after(0, self._done, out,o,f,e,em))
        threading.Thread(target=worker, daemon=True).start()

    def _update_stats(self, done, total, ok, fail, err):
        if total > 0:
            self._pbar_value = (done / total) * 100
            self._redraw_pbar()
        self.stat_values["ok"].config(text=str(ok))
        self.stat_values["fail"].config(text=str(fail))
        self.stat_values["err"].config(text=str(err))
        self.stat_values["tot"].config(text=str(total))

    def _log_msg(self, msg, tag="info"):
        self._log.insert("end", f"{msg}\n", tag)
        self._log.see("end")

    def _done(self, output, ok, fail, err, em=None):
        self._running = False
        self._pbar_value = 100
        self._redraw_pbar()
        if em:
            self._log_msg(f"❌ FATAL ERROR: {em}", "error")
            messagebox.showerror("Error", em)
            return
        self._log_msg(f"━━━━━━━━━━━━━━━━━━━━━━━━━", "info")
        self._log_msg(f"✓ SUCCESS: {ok}  |  ✗ UNAVAIL: {fail}  |  ⚠ ERROR: {err}", "success")
        self._log_msg(f"━━━━━━━━━━━━━━━━━━━━━━━━━", "info")
        messagebox.showinfo("Complete", f"✓ Success: {ok}\n✗ Unavailable: {fail}\n⚠ Errors: {err}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
